from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from src.models.database import db, Charge, ChargeItem, ChargeFees
from src.services.charge_calculator import ChargeCalculatorService
import os
import tempfile
from datetime import datetime

class DebtSpreadsheetGenerator:
    """Gerador de planilhas de débito em PDF e Excel"""
    
    def __init__(self):
        self.temp_dir = tempfile.gettempdir()
        self.styles = getSampleStyleSheet()
        
    def generate_pdf(self, charge_id):
        """Gera planilha do débito em PDF"""
        try:
            # Buscar dados da cobrança
            calculator = ChargeCalculatorService()
            data = calculator.generate_debt_spreadsheet(charge_id)
            
            charge = data['charge']
            
            # Criar arquivo PDF
            filename = f"planilha_debito_{charge['chargeCode']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            filepath = os.path.join(self.temp_dir, filename)
            
            doc = SimpleDocTemplate(filepath, pagesize=A4)
            story = []
            
            # Título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=self.styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Center
            )
            
            story.append(Paragraph("PLANILHA DE CÁLCULO DE DÉBITO", title_style))
            story.append(Spacer(1, 20))
            
            # Informações da cobrança
            info_data = [
                ['Código da Cobrança:', charge['chargeCode']],
                ['Cliente:', charge['client']['person']['name'] if charge.get('client') else 'N/A'],
                ['Devedor:', charge['debtor']['name'] if charge.get('debtor') else 'N/A'],
                ['Data de Vencimento:', charge['dueDate']],
                ['Data do Cálculo:', data['calculation_date']],
                ['Período de Referência:', charge.get('referencePeriod', 'N/A')]
            ]
            
            info_table = Table(info_data, colWidths=[2*inch, 4*inch])
            info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('BACKGROUND', (1, 0), (1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(info_table)
            story.append(Spacer(1, 30))
            
            # Seção 1: Débito Principal
            if data['principal_items']:
                story.append(Paragraph("1. DÉBITO PRINCIPAL", self.styles['Heading2']))
                story.append(Spacer(1, 10))
                
                principal_data = [
                    ['Data Venc.', 'Discriminação', 'Valor Nominal', 'Correção', 'Juros', 'Multa', 'Subtotal']
                ]
                
                for item in data['principal_items']:
                    principal_data.append([
                        item['dueDate'],
                        item['description'][:30] + '...' if len(item['description']) > 30 else item['description'],
                        f"R$ {item['nominalAmount']:,.2f}",
                        f"R$ {item['monetaryCorrection']:,.2f}",
                        f"R$ {item['interestAmount']:,.2f}",
                        f"R$ {item['fineAmount']:,.2f}",
                        f"R$ {item['subtotal']:,.2f}"
                    ])
                
                # Linha de total
                principal_data.append([
                    '', 'TOTAL DÉBITO PRINCIPAL', '', '', '', '',
                    f"R$ {data['totals']['principal_amount']:,.2f}"
                ])
                
                principal_table = Table(principal_data, colWidths=[0.8*inch, 2*inch, 1*inch, 0.8*inch, 0.8*inch, 0.8*inch, 1*inch])
                principal_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
                    ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(principal_table)
                story.append(Spacer(1, 20))
            
            # Seção 2: Despesas de Cobrança
            if data['expense_items']:
                story.append(Paragraph("2. DESPESAS DE COBRANÇA", self.styles['Heading2']))
                story.append(Spacer(1, 10))
                
                expense_data = [
                    ['Data Venc.', 'Discriminação', 'Valor Nominal', 'Correção', 'Juros', 'Subtotal']
                ]
                
                for item in data['expense_items']:
                    expense_data.append([
                        item['dueDate'],
                        item['description'][:30] + '...' if len(item['description']) > 30 else item['description'],
                        f"R$ {item['nominalAmount']:,.2f}",
                        f"R$ {item['monetaryCorrection']:,.2f}",
                        f"R$ {item['interestAmount']:,.2f}",
                        f"R$ {item['subtotal']:,.2f}"
                    ])
                
                # Linha de total
                expense_data.append([
                    '', 'TOTAL DESPESAS', '', '', '',
                    f"R$ {data['totals']['expenses_amount']:,.2f}"
                ])
                
                expense_table = Table(expense_data, colWidths=[0.8*inch, 2.2*inch, 1*inch, 0.8*inch, 0.8*inch, 1*inch])
                expense_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('BACKGROUND', (0, -1), (-1, -1), colors.lightgreen),
                    ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(expense_table)
                story.append(Spacer(1, 20))
            
            # Seção 3: Honorários
            story.append(Paragraph("3. HONORÁRIOS", self.styles['Heading2']))
            story.append(Spacer(1, 10))
            
            fees_data = [
                ['Tipo', 'Base de Cálculo', 'Percentual/Valor', 'Valor Calculado']
            ]
            
            if data['totals']['extrajudicial_fees'] > 0:
                fees_data.append([
                    'Honorários Extrajudiciais',
                    f"R$ {data['totals']['principal_amount']:,.2f}",
                    'Conforme contrato',
                    f"R$ {data['totals']['extrajudicial_fees']:,.2f}"
                ])
            
            if data['totals']['execution_fees'] > 0:
                fees_data.append([
                    'Honorários de Execução',
                    f"R$ {data['totals']['principal_amount']:,.2f}",
                    'Conforme contrato',
                    f"R$ {data['totals']['execution_fees']:,.2f}"
                ])
            
            # Total honorários
            total_fees = data['totals']['extrajudicial_fees'] + data['totals']['execution_fees']
            fees_data.append([
                'TOTAL HONORÁRIOS', '', '', f"R$ {total_fees:,.2f}"
            ])
            
            fees_table = Table(fees_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
            fees_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightyellow),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(fees_table)
            story.append(Spacer(1, 20))
            
            # Seção 4: Multa Art. 523 CPC
            if data['totals']['art_523_fine'] > 0:
                story.append(Paragraph("4. MULTA ART. 523 CPC", self.styles['Heading2']))
                story.append(Spacer(1, 10))
                
                fine_data = [
                    ['Base de Cálculo', 'Percentual', 'Valor da Multa'],
                    [
                        f"R$ {data['totals']['subtotal_with_fees']:,.2f}",
                        '10%',
                        f"R$ {data['totals']['art_523_fine']:,.2f}"
                    ]
                ]
                
                fine_table = Table(fine_data, colWidths=[2*inch, 1.5*inch, 2*inch])
                fine_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ('BACKGROUND', (0, 1), (-1, 1), colors.lightcoral),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(fine_table)
                story.append(Spacer(1, 30))
            
            # Resumo Final
            story.append(Paragraph("RESUMO PARA PRESTAÇÃO DE CONTAS", self.styles['Heading2']))
            story.append(Spacer(1, 10))
            
            summary_data = [
                ['Destinatário', 'Valor', 'Percentual'],
                [
                    'Cliente (Principal + Despesas)',
                    f"R$ {data['breakdown']['client_amount']:,.2f}",
                    f"{data['percentages']['client_percentage']:.2f}%"
                ],
                [
                    'Advogado (Honorários)',
                    f"R$ {data['breakdown']['lawyer_amount']:,.2f}",
                    f"{data['percentages']['lawyer_percentage']:.2f}%"
                ],
                [
                    'Tribunal (Multa Art. 523)',
                    f"R$ {data['breakdown']['court_amount']:,.2f}",
                    f"{data['percentages']['court_percentage']:.2f}%"
                ],
                [
                    'TOTAL GERAL',
                    f"R$ {data['breakdown']['total_amount']:,.2f}",
                    '100.00%'
                ]
            ]
            
            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch, 1.5*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightsteelblue),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(summary_table)
            
            # Gerar PDF
            doc.build(story)
            
            return filepath
            
        except Exception as e:
            raise Exception(f"Erro ao gerar PDF: {str(e)}")
    
    def generate_excel(self, charge_id):
        """Gera planilha do débito em Excel"""
        try:
            # Buscar dados da cobrança
            calculator = ChargeCalculatorService()
            data = calculator.generate_debt_spreadsheet(charge_id)
            
            charge = data['charge']
            
            # Criar arquivo Excel
            filename = f"planilha_debito_{charge['chargeCode']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.temp_dir, filename)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Planilha de Débito"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Título
            ws.merge_cells('A1:G1')
            ws['A1'] = "PLANILHA DE CÁLCULO DE DÉBITO"
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Informações da cobrança
            row = 3
            info_data = [
                ['Código da Cobrança:', charge['chargeCode']],
                ['Cliente:', charge['client']['person']['name'] if charge.get('client') else 'N/A'],
                ['Devedor:', charge['debtor']['name'] if charge.get('debtor') else 'N/A'],
                ['Data de Vencimento:', charge['dueDate']],
                ['Data do Cálculo:', data['calculation_date']],
                ['Período de Referência:', charge.get('referencePeriod', 'N/A')]
            ]
            
            for info in info_data:
                ws[f'A{row}'] = info[0]
                ws[f'B{row}'] = info[1]
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
            
            row += 2
            
            # Débito Principal
            if data['principal_items']:
                ws[f'A{row}'] = "1. DÉBITO PRINCIPAL"
                ws[f'A{row}'].font = Font(bold=True, size=14)
                row += 1
                
                headers = ['Data Venc.', 'Discriminação', 'Valor Nominal', 'Correção', 'Juros', 'Multa', 'Subtotal']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                
                row += 1
                
                for item in data['principal_items']:
                    ws[f'A{row}'] = item['dueDate']
                    ws[f'B{row}'] = item['description']
                    ws[f'C{row}'] = item['nominalAmount']
                    ws[f'D{row}'] = item['monetaryCorrection']
                    ws[f'E{row}'] = item['interestAmount']
                    ws[f'F{row}'] = item['fineAmount']
                    ws[f'G{row}'] = item['subtotal']
                    
                    # Aplicar bordas e formatação de moeda
                    for col in range(1, 8):
                        cell = ws.cell(row=row, column=col)
                        cell.border = border
                        if col >= 3:  # Colunas de valor
                            cell.number_format = 'R$ #,##0.00'
                    
                    row += 1
                
                # Total
                ws[f'B{row}'] = "TOTAL DÉBITO PRINCIPAL"
                ws[f'G{row}'] = data['totals']['principal_amount']
                ws[f'B{row}'].font = Font(bold=True)
                ws[f'G{row}'].font = Font(bold=True)
                ws[f'G{row}'].number_format = 'R$ #,##0.00'
                
                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = border
                
                row += 3
            
            # Resumo Final
            ws[f'A{row}'] = "RESUMO PARA PRESTAÇÃO DE CONTAS"
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 1
            
            summary_headers = ['Destinatário', 'Valor', 'Percentual']
            for col, header in enumerate(summary_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            row += 1
            
            summary_data = [
                ['Cliente (Principal + Despesas)', data['breakdown']['client_amount'], data['percentages']['client_percentage']],
                ['Advogado (Honorários)', data['breakdown']['lawyer_amount'], data['percentages']['lawyer_percentage']],
                ['Tribunal (Multa Art. 523)', data['breakdown']['court_amount'], data['percentages']['court_percentage']],
                ['TOTAL GERAL', data['breakdown']['total_amount'], 100.00]
            ]
            
            for summary in summary_data:
                ws[f'A{row}'] = summary[0]
                ws[f'B{row}'] = summary[1]
                ws[f'C{row}'] = f"{summary[2]:.2f}%"
                
                ws[f'B{row}'].number_format = 'R$ #,##0.00'
                
                for col in range(1, 4):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    if row == row:  # Última linha
                        cell.font = Font(bold=True)
                
                row += 1
            
            # Ajustar largura das colunas
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            
            # Salvar arquivo
            wb.save(filepath)
            
            return filepath
            
        except Exception as e:
            raise Exception(f"Erro ao gerar Excel: {str(e)}")

